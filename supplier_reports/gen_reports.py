import logging
from openpyxl import load_workbook
import copy
import contextlib
import os
import unicodecsv as csv
from werkzeug.datastructures import FileStorage
import functools,StringIO
from collections import OrderedDict
from python_script_common import exception_context_extra_info

#####################################################################
#Globals
_logger = logging.getLogger(__name__)


#####################################################################
# move to env utiliy funcs


def read_xlsx_sheet(file_path, sheet_name):
    wb = load_workbook(file_path)
    assert sheet_name in wb, "no such sheet name {}".format(sheet_name)
    sheet = wb[sheet_name]
    row_count = sheet.max_row
    column_count = sheet.max_column

    header = []

    for col in range(1, column_count+1):
        header.append(sheet.cell(1, col).value)

    data = []
    for row in range(2, row_count+1):
        rdict = OrderedDict([(header[col-1],sheet.cell(row, col).value) for col in range(1, column_count+1)])
        data.append(rdict)
    return data

#####################################################################
# errors


class EmptyReportError(Exception):
    pass


class LookupKeyError(Exception):
    pass


class PrimaryKeyError(Exception):
    pass


class ParsingError(Exception):
    pass


class GroupingError(Exception):
    pass


class EmptyValueError(Exception):
    pass

#####################################################################
# general schema validators


class LookupDict(dict):
    def __init__(self,list_of_dicts, fields_in_index):
        super(LookupDict, self).__init__()
        self.fields_in_index=fields_in_index
        self.fields=list_of_dicts[0].keys()
        assert set(fields_in_index).issubset(self.fields)
        for rowdict in list_of_dicts:
            index_str = self.create_index_string(rowdict)
            self[index_str]=rowdict

    def create_index_string(self, rdict):
        index_values = [rdict[key] for key in self.fields_in_index]
        index_string = "_".join(map(str, index_values))
        return index_string

    def get_matching_dict_for(self, rdict):
        try:
            idx = self.create_index_string(rdict)
            return self[self.create_index_string(rdict)]
        except KeyError as e:
            raise LookupKeyError("unable to lookup key '{}'".format(idx))


class SimpleSchemaValidator(object):
    def __init__(self, schema):
        self.schema = schema
        self.is_table_context = False
        self.grouped_info = {}
        self.grouped_rows=[]
        self._current_group_context = None
        self.primary_keys = []

    def validate(self, data_dict, post_process=False):
        if "primary_keys" in self.schema:
            self.validate_primary_keys(data_dict)
        if "not_empty" in self.schema:
            self.validate_not_empty(data_dict)
        if post_process:
            if "fill_missing" in self.schema:
                data_dict = self.fill_missing(data_dict)

        return data_dict

    def validate_not_empty(self, data_dict):
        assert isinstance(data_dict, dict)
        for key in self.schema["not_empty"]:
            if not data_dict[key]:
                raise EmptyValueError("{} in {} has no value".format(key, data_dict))

    def validate_primary_keys(self, data_dict):
        assert isinstance(data_dict, dict)
        for key in self.schema["primary_keys"]:
            if key not in data_dict:
                raise PrimaryKeyError("key_not_in_data: {}".format(key))
            if not data_dict[key]:
                raise PrimaryKeyError("key_should_not_be_empty: {}".format(data_dict[key]))
            if isinstance(data_dict[key], basestring) and data_dict[key] != data_dict[key].rstrip():
                data_dict[key]=data_dict[key].rstrip()
                #raise PrimaryKeyError("key_should_not_end_with_hidden_chars: {}".format(data_dict[key]))

        if self.is_table_context:
            pk_vector = [data_dict[k] for k in self.schema["primary_keys"]]
            if pk_vector in self.primary_keys:
                raise PrimaryKeyError("duplicate_primary_key {}".format(pk_vector))
            else:
                self.primary_keys.append(pk_vector)

    def fill_missing(self, data_dict):
        assert isinstance(data_dict, dict)
        new_data = copy.deepcopy(data_dict)
        for missing_defs in self.schema["fill_missing"]:
            target_field, from_fields = missing_defs["field"], missing_defs["from"]
            if not new_data[target_field]:
                try:
                    new_data[target_field] = next(data_dict[key] for key in from_fields if data_dict[key])
                except StopIteration as e:
                    pass
        return new_data

    def populate_cached_rows_with_group_data(self):
        for rowd in self.grouped_rows:
            rowd.update(self.grouped_info[self._current_group_context])

    def generate_grouped_data(self, rowdict):
        """ while in grouping context update grouped info, retaining grouping rows
            after grouping context changes, update all cached rows
        :param rowdict:
        :return:
        """
        #import pdb;pdb.set_trace()
        grouped_field_name = self.schema["fill_grouped"]["on"]
        group_key = rowdict[grouped_field_name]
        by_field_names = self.schema["fill_grouped"]["by"]

        if self._current_group_context == group_key: #inside grouping context
            for k in by_field_names:
                if rowdict[k]:
                    self.grouped_info[group_key][k]=rowdict[k]

            #self.grouped_info[group_key] = {k: rowdict[k] for k in by_field_names if rowdict[k] is not None}
            self.grouped_rows.append(rowdict)
        else:
            if group_key in self.grouped_info:
                raise GroupingError("{} was already grouped once".format(group_key))
            else: #grouping context changes
                self.populate_cached_rows_with_group_data()
                self._current_group_context = group_key
                self.grouped_info[group_key] = {k: rowdict[k] for k in by_field_names if rowdict[k] is not None}
                self.grouped_rows=[rowdict]

    @contextlib.contextmanager
    def table_context(self):
        try:
            self.grouped_info = {}
            self.is_table_context = True
            self._current_group_context = None
            self.primary_keys = []
            yield
        finally:
            if "fill_grouped" in self.schema:
                self.populate_cached_rows_with_group_data()
            self.is_table_context = False

    def validate_table(self, list_of_dicts, post_process=False):
        list_of_dicts_result=[]
        with self.table_context():
            for i,rowdict in enumerate(list_of_dicts):
                row = i + 1
                with exception_context_extra_info("error in row {}".format(row), "row=<{}>".format(rowdict)):
                    _logger.debug("processing row {}".format(row))
                    processed_rowdict = self.validate(rowdict, post_process=post_process)
                    if post_process:
                        if "fill_grouped" in self.schema:
                            self.generate_grouped_data(processed_rowdict)
                    list_of_dicts_result.append(processed_rowdict)
        return list_of_dicts_result

    def export_fields(self, list_of_dicts, lookup_dict=None):
        local_fields = [f for f in self.schema["report_fields"] if f in list_of_dicts[0].keys()]

        if "new_fields" in self.schema:
            new_fields = self.schema["new_fields"]
        else:
            new_fields = []

        if lookup_dict:
            assert isinstance(lookup_dict, LookupDict)
            expected_looked_up_fields = set(self.schema["report_fields"])-set(new_fields)-set(local_fields)
            diff = set(expected_looked_up_fields)-set(lookup_dict.fields)
            assert not diff, "lookup dict missing keys {}".format(diff)

        report_table = []
        match_row_key = self.schema["match_row_key"]
        match_row_value = self.schema["match_row_value"]
        for i, rowdict in enumerate(list_of_dicts):
            _logger.debug("processing row {}".format(i+1))
            with exception_context_extra_info("error in row {}".format(i+1)):
                if match_row_key in rowdict:
                    if match_row_value == rowdict[match_row_key]:
                        # initiate report with local fields
                        report_row={k:v for k,v in rowdict.items() if k in local_fields}
                        # add empty fields
                        report_row.update({k:None for k in new_fields})
                        if lookup_dict:
                            # add lookup fields
                            match = lookup_dict.get_matching_dict_for(rowdict)
                            report_row.update({k:match[k] for k in expected_looked_up_fields})

                        report_table.append(report_row)

        if not report_table:
            _logger.warning("No data for report:'{}'".format(self.schema))
        return report_table


class CsvReport(dict):
    def __repr__(self):
        return "<Report for schema:{}>".format(self.schema["description"])

    def __init__(self, schema, file=None, prefix=None):
        super(CsvReport,self).__init__()
        self.schema = schema
        self.file = file
        self.prefix = prefix

    def save(self, directory):
        report_path = os.path.join(directory, self.get_report_file_name())
        _logger.info("saving {}".format(report_path))
        with open(report_path, "w") as f:
            f.write(self.file.getvalue())

    def get_report_file_name(self):
        report_filename = self.schema["match_row_value"].replace(" ", "_")
        if self.prefix:
            report_filename = "{}-{}.csv".format(self.prefix, report_filename)
        return report_filename


class SchemaDescription(dict):
    def __repr__(self):
        return self["description"]

    def __init__(self, data_dict):
        super(SchemaDescription,self).__init__()
        self.update(data_dict)


schema_product_list = SchemaDescription({
    "description": "data schema product_list",
    'primary_keys': ["Vendor", "Lineitem name"],
                     'not_empty': ["Product link"]
                    }
)

schema_export_orders = SchemaDescription({
    "description": "data schema export_orders",
    'fill_missing': [{"field":'Shipping Phone', "from":['Phone','Billing Phone']}],
    'fill_grouped': {  'on': 'Name',
                        'by':  ["Shipping Name","Shipping Street",
                                "Shipping Address1","Shipping Address2","Shipping Company",
                                "Shipping City","Shipping Zip","Shipping Province","Shipping Country",
                                "Shipping Phone","Notes"]
                    }
       }
)

report_schema_mr_art_painting = SchemaDescription({
    "description":"report schema Mr art Painting store",
    'match_row_key':'Vendor',
    'match_row_value':"Mr Art Painting store",
    'report_fields':  ["Name", "Lineitem name", "Variant", "Lineitem quantity", "Product link", "Size",
                "Frame option", "Shipping Name", "Shipping Street", "Shipping Address1",
                "Shipping Address2", "Shipping Company", "Shipping City", "Shipping Zip",
                "Shipping Province", "Shipping Country", "Shipping Phone"]


})

report_schema_mr_zhen = SchemaDescription({
    "description":"report schema mr Zhen",
    'match_row_key': 'Vendor',
    'match_row_value': "Zhen",
    'report_fields':  ["Name", "Lineitem name", "Created at", "Financial status", "Fulfillment Status", "Internal note", "Marketplace",
                "Supplier", "Product link", "Lineitem sku", "Variant", "Size", "Color", "Lineitem quantity",
                "Price", "Shipping Name", "Shipping Street", "Shipping Address1","Tracking Number",
                "Shipping Address2", "Shipping Company", "Shipping City", "Shipping Zip", "Shipping Province", 
                "Shipping Country", "Shipping Phone"],
    'new_fields': ["Financial status","Internal note","Marketplace","Supplier","Tracking Number","Price"]
})




#####################################################################


def import_product_list(file_path):
    sheet="Product list"
    table_data = read_xlsx_sheet(file_path, sheet)
    sv = SimpleSchemaValidator(schema_product_list)
    _logger.info("validating '{}'".format(sheet))
    return sv.validate_table(table_data)


def import_export_orders(file_path):
    sheet="Export orders"
    table_data = read_xlsx_sheet(file_path, sheet)
    sv = SimpleSchemaValidator(schema_export_orders)
    _logger.info("validating '{}'".format(sheet))
    return sv.validate_table(table_data,post_process=True)


def get_csv_report(list_of_dicts, field_order):
    diff = set(field_order) - set(list_of_dicts[0].keys())
    assert not diff, "export to csv missing fields: {}".format(diff)
    file = StringIO.StringIO()
    dict_writer = csv.DictWriter(file, field_order)
    dict_writer.writeheader()
    dict_writer.writerows(list_of_dicts)
    return file


def get_file_name(filepath):
    path, file_name = os.path.split(filepath)
    file_name, ext = os.path.splitext(file_name)
    return file_name


def gen_reports(filepath_or_filestorage):
    if isinstance(filepath_or_filestorage, FileStorage):
        import_export_orders_file_name = filepath_or_filestorage.filename
    else:
        import_export_orders_file_name = get_file_name(filepath_or_filestorage)

    import_export_orders_lod = import_export_orders(filepath_or_filestorage)
    product_list_lod = import_product_list(filepath_or_filestorage)
    lookupd = LookupDict(product_list_lod, schema_product_list["primary_keys"])
    for schema in (globals()[k] for k in globals() if k.startswith("report_schema")):
        _logger.info("processing data for report: {}".format(schema))
        sv = SimpleSchemaValidator(schema)
        try:
            report_lod = sv.export_fields(import_export_orders_lod, lookupd)
        except LookupKeyError as e:
            _logger.error("unable to locate product list key")
            raise
        if report_lod:
            field_order = schema["report_fields"]
            report_file = get_csv_report(report_lod, field_order)
            _logger.info("generated filestorage csv report")
            report = CsvReport(schema, report_file, prefix=import_export_orders_file_name)
        else:
            report = None

        yield report







