# -*- coding: utf-8 -*-

import pytest
from openpyxl import Workbook
from openpyxl.compat import range
from supplier_reports.gen_reports import read_xlsx_sheet, SimpleSchemaValidator,GroupingError
from supplier_reports import gen_reports
import tempfile,os
import logging
import copy
import contextlib
import shutil
import atexit

logging.basicConfig(level=logging.DEBUG)
_logger = logging.getLogger()


def create_xlsx_sheet(filename, title, data):
    wb = Workbook()
    ws = wb.create_sheet(title=title)
    for j,header_value in enumerate(data[0].keys()):
        ws.cell(column=j+1, row=1, value=header_value)
    for i, rowdict in enumerate(data):
        for j, value in enumerate(rowdict.values()):
            ws.cell(column=j+1, row=i+2, value=value)
    wb.save(filename=filename)
    return wb


@contextlib.contextmanager
def tempdir_context():
    dir_path = tempfile.mkdtemp()
    try:
        yield dir_path
    finally:
        shutil.rmtree(dir_path)


def test_read_xlsx_sheet():
    data=[]
    for x in range(5):
        data.append(dict(name=x,age=x,sex=x))
    xls_file_name="test_sheet.xlsx"
    with tempdir_context() as tmpdirname:
        file_path=os.path.join(tmpdirname,xls_file_name)
        wb = create_xlsx_sheet(file_path, "test1", data)
        res=read_xlsx_sheet(file_path, "test1")
        assert data == res
        wb.close()


@pytest.fixture
def valid_grouped_first_line():
    schema = {'fill_grouped': {'on': 'ord_num',
                                'by': ['user_name','user_phone']
                            }
            }
    input_table = [{'ord_num': 1, 'user_name': 1, 'user_phone': 1, 'supplier':1},
                  {'ord_num': 1, 'user_name': None, 'user_phone': None, 'supplier':2},
                  {'ord_num': 2, 'user_name': 2, 'user_phone': 2, 'supplier':3},
                  {'ord_num': 2, 'user_name': None, 'user_phone': None, 'supplier':4}]

    expected_table = [{'ord_num': 1, 'user_name': 1, 'user_phone': 1, 'supplier':1},
                  {'ord_num': 1, 'user_name': 1, 'user_phone': 1, 'supplier':2},
                  {'ord_num': 2, 'user_name': 2, 'user_phone': 2, 'supplier':3},
                  {'ord_num': 2, 'user_name': 2, 'user_phone': 2, 'supplier':4}]

    return schema, input_table, expected_table


def test_grouping_first_line(valid_grouped_first_line):
    schema, input_table, expected_table = valid_grouped_first_line

    _logger.info("test for grouping when grouped data is in the first line of the group")
    sv = SimpleSchemaValidator(schema)
    grouped_table = sv.validate_table(input_table, post_process=True)
    assert grouped_table == expected_table


def test_grouping_error(valid_grouped_first_line):
    schema, input_table, expected_table = valid_grouped_first_line
    invalid_data = copy.copy(input_table)
    invalid_data.append(input_table[0])

    _logger.info("test for repeating row table, expected to fail")
    sv = SimpleSchemaValidator(schema)

    with pytest.raises(GroupingError) as e:
        sv.validate_table(invalid_data, post_process=True)


def test_grouping_last_line(valid_grouped_first_line):
    """ reversing row 1&2, 3&4 so that grouping data is always on the last line

    :param valid_grouped_first_line:
    :return:
    """
    schema, valid_data, expected_table = valid_grouped_first_line

    line1,line2,line3,line4 = valid_data
    input_table_grouping_data_last_line = [line2, line1, line4, line3]

    line1_exp, line2_exp, line3_exp, line4_exp = expected_table
    expected_table_new=[line2_exp, line1_exp,line4_exp,line3_exp]

    _logger.info("test for grouping when grouped data is in the last line of the group")
    sv = SimpleSchemaValidator(schema)

    grouped_table = sv.validate_table(input_table_grouping_data_last_line, post_process=True)
    assert grouped_table == expected_table_new


def test_fill_missing():
    schema = {'fill_missing': [{'field': 'a','from': ['b','c']},]
              }

    data1 = {'a': None, 'b': 2, 'c': 3}
    data2 = {'a': 1, 'b': 2, 'c': 3}
    data3 = {'a': None, 'b': None, 'c': None}
    data4 = {'a': None, 'b': None, 'c': 3}

    sv = SimpleSchemaValidator(schema)

    d1 = sv.validate(data1, post_process=True)
    assert d1["a"] == 2
    assert d1["b"] == data1["b"] and d1["c"] == data1["c"], "other data should not change"

    d2 = sv.validate(data2, post_process=True)
    assert d2["a"] == 1
    assert d1["b"] == data1["b"] and d1["c"] == data1["c"], "other data should not change"

    d3 = sv.validate(data3, post_process=True)
    assert d3["a"] is None
    assert d1["b"] == data1["b"] and d1["c"] == data1["c"], "other data should not change"

    d4 = sv.validate(data4, post_process=True)
    assert d4["a"] == 3
    assert d1["b"] == data1["b"] and d1["c"] == data1["c"], "other data should not change"


def test_primary_keys():
    schema = {'primary_keys': ['a']
              }

    data1 = [{'a': 1, 'b': 1, 'c': 1},
              {'a': 2, 'b': 2, 'c': 2},
              {'a': 3, 'b': 3, 'c': 3},
              {'a': 4, 'b': 4, 'c': 4}]

    sv = SimpleSchemaValidator(schema)

    _logger.info("validate good pk")
    sv.validate_table(data1, post_process=True)

    _logger.info("validate duplicate pk")
    data2 = copy.deepcopy(data1)
    data2.append(data1[0])
    with pytest.raises(gen_reports.PrimaryKeyError) as e:
        sv.validate_table(data2, post_process=True)

    _logger.info("validate duplicate pk")
    data3 = copy.deepcopy(data1)
    data3.append({'a': None, 'b': None, 'c': None})
    with pytest.raises(gen_reports.PrimaryKeyError) as e:
        sv.validate_table(data3, post_process=True)


def test_not_empty():
    schema = {'not_empty': ['a']
              }

    data1 = {'a': 1, 'b': 2, 'c': 3}
    data2 = {'a': None, 'b': 2, 'c': 3}
    data3 = {'a': "", 'b': 2, 'c': 3}

    sv = SimpleSchemaValidator(schema)

    _logger.info("non empty ok")
    sv.validate(data1)

    _logger.info("non empty catches None")
    with pytest.raises(gen_reports.EmptyValueError):
        sv.validate(data2)

    _logger.info("non empty catches empty string")
    with pytest.raises(gen_reports.EmptyValueError):
        sv.validate(data3)


@pytest.fixture
def valid_export_data():
    valid_data = [{'a': 1, 'b': 1, 'c': 1},
                  {'a': 1, 'b': None, 'c': None},
                  {'a': 2, 'b': 2, 'c': 2},
                  {'a': 2, 'b': None, 'c': None}]
    return valid_data


def test_export_local_match(valid_export_data):
    report_schema = {
        'match_row_key': 'a',
        'match_row_value': 1,
        'report_fields': ["a","b"]
    }
    sv = SimpleSchemaValidator(report_schema)
    report = sv.export_fields(valid_export_data)
    expected = [{'a': 1, 'b': 1},
                  {'a': 1, 'b': None, }]
    assert report == expected, "report:\n'{}' does not match expected:\n'{}'".format(report, expected)


def test_export_local_no_match_key(valid_export_data):
    _logger.info("test no local match key no lookup")
    report_schema = {
        'match_row_key': 'd',
        'match_row_value': 1,
        'report_fields': ["a","b"]
    }
    sv = SimpleSchemaValidator(report_schema)
    report = sv.export_fields(valid_export_data)
    assert report == [], "we should have an empty report"


def test_export_local_no_match_value(valid_export_data):
    _logger.info("test no local match key no lookup")
    report_schema = {
        'match_row_key': 'a',
        'match_row_value': 11,
        'report_fields': ["a", "b"]
    }
    sv = SimpleSchemaValidator(report_schema)
    report = sv.export_fields(valid_export_data)
    assert report == [], "we should have an empty report"


def test_export_with_lookup_match(valid_export_data):
    _logger.info("test no local match key no lookup")
    report_schema = {
        'match_row_key': 'a',
        'match_row_value': 1,
        'report_fields': ["a", "b", "e"]
    }

    lookup_lod = [
        {'a':1,'e': 2},
        {'a':2,'e': 3},
        {'a':3,'e': 4},
    ]

    sv = SimpleSchemaValidator(report_schema)
    lookupd = gen_reports.LookupDict(lookup_lod, ['a'])
    report = sv.export_fields(valid_export_data, lookupd)
    expected = [{'a': 1, 'b': 1, 'e': 2},
                  {'a': 1, 'b': None, 'e': 2}]
    assert report == expected, "report:\n'{}' does not match expected:\n'{}'".format(report,expected)


def test_export_with_lookup_match_error(valid_export_data):
    _logger.info("test no local match key no lookup")
    report_schema = {
        'match_row_key': 'a',
        'match_row_value': 1,
        'report_fields': ["a", "b", "e"]
    }

    lookup_lod = [
        {'a':2,'e': 3},
        {'a':3,'e': 4},
    ]

    sv = SimpleSchemaValidator(report_schema)
    lookupd = gen_reports.LookupDict(lookup_lod, ['a'])
    with pytest.raises(gen_reports.LookupKeyError):
        sv.export_fields(valid_export_data, lookupd)
